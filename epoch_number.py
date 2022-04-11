import requests
import xlsxwriter
from openpyxl import load_workbook
import time

def get_cell_data(sheet,row,col):
    wb = load_workbook(r"C:\Users\kajet\OneDrive\Pulpit\szablon.xlsx")

    sheets = wb.sheetnames
    Sheet1 = wb[sheets[sheet]]
    cell_data = Sheet1.cell(row=row, column=col).value
    int(cell_data or 0)
    #int(0 if cell_data is None else cell_data)
    return int(0 if cell_data is None else cell_data)


def add_data(sheet,row,col,value):
    wb = load_workbook(r"C:\Users\kajet\OneDrive\Pulpit\szablon.xlsx")

    sheets = wb.sheetnames
    Sheet1 = wb[sheets[sheet]]
    Sheet1.cell(row =row, column = col).value = value

    wb.save(r"C:\Users\kajet\OneDrive\Pulpit\szablon.xlsx")

url = 'https://api.idena.io/api/Epochs?limit=1'

identity_table = [
'0x344a09Ec5B9B5deBC5a889837C50c66C8a78f04D',
'0xD39aafE2D3814cF6f3afaE9f446192b92688eA9A',
'0xf462697f2589edd6958acd2126b0a05130222414',
'0x699c0384082c04B4DA95217db0D2052f4802CEC3',
'0x5FB3eBcD91e5EF04fE783872Ffa6Ac3178B6f003',
'0xF6b5d45BeB1B651B43E4F6862D865d98564C83DB',
'0xbE15EA49A37F853d5bB3514717D78daab137D807',
'0x0392C1F57a61A40eE5d14f7B11760f7055f95468',
'0x1df4a39030Bcd1d2ad58713cA18FD97c8e848CC0',
'0x810F20Bd6A29F8Ee57E4f1B90314e507e2bD96Bc',
'0xb9E6aE2A163Ed74A4A2560a3Ff4Cef302bAbbe8e',
'0xF4F010AD8a781BA8A574763A0853881F223550F2',
'0x4Af9d0f5a405b0db32970fd88e84fD72F8BE0C63',
'0x024392becDb3E86c45fBC318d99E296D5fde08a9',
'0x8BD34f9E6b768B03736fB586AbdBDC1f0CCEc3b6',
'0x2e5f5ff80AB45e6Eb8D201404e265D72564ED026',
'0xc34D15574EC0B650cC19ffCF3489A6Ae0b3f25C6',
'0x4466472097a5924b6284bdc5d530375031c3D5fa',
'0x204e310f79659FaD31D6c7D129B852905405794B',
'0x3b59F99Ae3393b121279111F7cB7475262171F40',
'0xC8fF46DdcF574b32d63A3D45b54bB3c6d8C265cE',
'0x1eFfe69fa5d4BcB0b2eF8bF6bB08a889617742E7',
'0x187a0720cc95FA8EE99F342e2CB31C670b5a98D1',
'0xd199d2A9f00AF81EE237380386EE0955762546D8'
]

def get_epoch_api_data():
    result = requests.get("https://api.idena.io/api/Epochs?limit=1")
    return result

def last_block_height():
    result = requests.get("https://api.idena.io/api/Block/Last")
    headers = result.text.split(',')[1]
    height = headers.split(':')[1]
    return height

def get_epoch_numb():
    result = get_epoch_api_data()
    headers = result.text.split(',')[0]
    epoch = headers.split(':')[2]
    return epoch

def get_result(address):
    result = requests.get("https://api.idena.io/api/Address/"+address)
    return result

def txCount(identity):
    result = get_result(identity)
    headers = result.text.split(',')[3]
    txCount = headers.split(':')[1]
    return txCount


def sendtx_address_to(block_height):
    url = "https://api.idena.io/api/Block/"+str(block_height)+"/Txs?limit=100"
    result = requests.get(url)
    count = 0
    #print(result.text)
    headers = result.text.split(',')
    for x in range(0,len(headers)):
        if headers[x] == '"type":"SendTx"':
            to = headers[x+3]
            amount = float(headers[x+4].split('"')[3]) #sendAmount
            address = to.split('"')[3] #to_address
            for y in range(0,len(identity_table)): # mozna brakeować po tym jak wejdzie w ifa i doda do excela dane
                if address == identity_table[y] and amount <=6:
                    add_data(0,3+y,3,int(float(get_cell_data(0,3+y,3))+1))
                    break
            count = count + 1
            print(address)
    if count == 0:
        print(" Brak sendów")

last_block_height =4197659 #duże litery w adresach nie wyłapuje

while True: #puścić od epochu 80, jeżeli block nie jest ostatnim z api to bez sleepa a jak juz dojdzie do ostatniego to sleep 15
    print(last_block_height,end='')
    sendtx_address_to(last_block_height)
    #time.sleep(15)
    last_block_height= last_block_height + 1
    if int(last_block_height) == 4267875:
        break

# sendtx_address_to(4267622)